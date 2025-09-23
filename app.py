import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, Font
from io import BytesIO

# --- PASSWORD PROTECTION ---
st.title("üîí RespIndNet Specimen Transfer Form (Virology)")
password = st.text_input("Enter Password:", type="password")

if password != "RIN123":  # <-- Change password here
    st.warning("Please enter the correct password.")
    st.stop()

# --- Load Google Sheets ---
sheet_id1 = "10DQqdiCM1bJEj14U6TO-obcdm_2r15wHOWpwlBdm1qY"
csv_url1 = f"https://docs.google.com/spreadsheets/d/{sheet_id1}/gviz/tq?tqx=out:csv"
df1 = pd.read_csv(csv_url1, on_bad_lines="skip")

if "dt_delivery" in df1.columns:
    df1["dt_delivery"] = pd.to_datetime(df1["dt_delivery"], errors="coerce")

sheet_id = "1Re-0aMf2Xok0CvcaSQPIsKhhyTXFF_Sx-xra3n5wiYg"
csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv"
df = pd.read_csv(csv_url, on_bad_lines="skip")

# Normalize
df.columns = df.columns.str.strip().str.lower()
df["submissiondate"] = pd.to_datetime(df["submissiondate"], errors="coerce")
df["submissiondate"] = df["submissiondate"].dt.tz_localize(None)

# Filter today‚Äôs data
today_str = pd.Timestamp.today().strftime("%Y-%m-%d")
df_today = df[(df["submissiondate"].dt.strftime("%Y-%m-%d") == today_str) & (df["sp_col"] == 1)].copy()

# Barcode ID
df_today["barcode_id"] = np.where(
    df_today["sample_scan"].notna() & (df_today["sample_scan"].astype(str).str.strip() != ""),
    df_today["sample_scan"],
    df_today["sample_scan_manually"]
)

# Sample type mapping
stype_mapping = {"B": "Blood", "R": "Respiratory swab", "D": "Diarrhoea", "S": "Stool (monthly)"}
df_today["stype"] = df_today["stype"].astype(str).str.strip().str.upper()
df_today["type_of_sample"] = df_today["stype"].map(stype_mapping)
df_today = df_today[df_today["type_of_sample"].isin(["Blood", "Respiratory swab"])]
df_today["sample_sequence"] = df_today.groupby("child_id").cumcount() + 1

# Merge delivery date
merged = df_today.merge(df1[["cid", "dt_delivery"]], left_on="child_id", right_on="cid", how="left")

# Age calculation
def calculate_age(delivery_date):
    if pd.isna(delivery_date): return ""
    today = datetime.today()
    years = today.year - delivery_date.year
    months = today.month - delivery_date.month
    days = today.day - delivery_date.day
    if days < 0: months -= 1
    if months < 0:
        years -= 1
        months += 12
    if years == 0:
        return f"{months} months"
    elif months == 0:
        return f"{years} year{'s' if years > 1 else ''}"
    else:
        return f"{years} year{'s' if years > 1 else ''} {months} months"

merged["AGE"] = merged["dt_delivery"].apply(calculate_age)
merged["VOLUME (FIELD)"] = np.where(merged["type_of_sample"] == "Blood", merged["sp_vol"].astype(str) + " ML", "")

# Final table
table = pd.DataFrame({
    "S. NO": range(1, len(merged) + 1),
    "BARCODE ID": merged["barcode_id"],
    "DATE (DD-MM-YYYY)": merged["submissiondate"].dt.strftime("%d-%m-%Y"),
    "TYPE OF SAMPLE": merged["type_of_sample"],
    "SAMPLE SEQUENCE (SAMPLE NUMBER PER INDIVIDUAL)": merged["sample_sequence"],
    "IND ID": merged.get("child_id", ""),
    "NAME": "Mrs. " + merged["mo_name"].astype(str) + "'s Baby",
    "AGE": merged["AGE"],
    "SAMPLE COLLECTION DATE & TIME": merged["submissiondate"],
    "VOLUME (FIELD)": merged["VOLUME (FIELD)"],
    "RECEIVED BY": "",
    "MEASURED VOLUME (VIROLOGY)": "",
    "REMARKS (LYSED/ LIPEMIC/ ICTERIC / SAMPLE SPILLAGE)": ""
})

st.subheader("üìã Generated Table")
st.dataframe(table)

# --- Download Excel ---
if len(table) > 0:
    today_str = datetime.today().strftime("%d-%m-%Y")
    excel_filename = f"{today_str}_RespIndNet_STF(Field to Virology).xlsx"

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Header row 1 (merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table.shape[1])
    cell = ws.cell(row=1, column=1, value="RespIndNet_Study Specimen Transfer Form (Virology)")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

    # Extra span row (row 2)
    spans = [
    "Sample Shipment Date and Time:",
    "Field Manager Sign/Initials:",
    "Virology Staff Sign/Initials:",
    "",   # empty span
    "To be filled by Virology"
]

# Split table width into 5 parts
col_split = [
    table.shape[1] // 5,
    table.shape[1] // 5,
    table.shape[1] // 5,
    table.shape[1] // 5,   # empty span width
    table.shape[1] - 4 * (table.shape[1] // 5)   # remaining for last span
]

start_col = 1
for i, val in enumerate(spans):
    end_col = start_col + col_split[i] - 1
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    cell = ws.cell(row=2, column=start_col, value=val if val else None)  # leave empty span blank
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # Apply border to merged range
    for row in range(2, 3):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border
    start_col = end_col + 1

    # Force row 2 height bigger so wrapped text is visible
    ws.row_dimensions[2].height = 30


    # Column headers (row 3)
    for j, col_name in enumerate(table.columns, 1):
        c = ws.cell(row=3, column=j, value=col_name)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # Data rows
    for i, row in table.iterrows():
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i+4, column=j, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

    # Save to memory for download
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="‚¨áÔ∏è Download Excel",
        data=buffer,
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

